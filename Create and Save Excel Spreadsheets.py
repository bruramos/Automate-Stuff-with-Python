import openpyxl, os

os.chdir('C:\\Users\\Bruna Ramos\\OneDrive - SMARTHIS CONSULTORIA EM TECNOLOGIA DA INFORMAÇÃO LTDA\\Documentos\\Automate Stuff with Python')

wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb['Sheet']

sheet['A1'] = 42
sheet['A2'] = 'Hello'

sheet2 = wb.create_sheet()

sheet3 = wb.create_sheet(index=0, title='Sheet3')

wb.save('Teste.xlsx')


